
import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import matplotlib.pyplot as plt

st.set_page_config(page_title="Resumen de Cajas y Picking", layout="wide")
st.title("📦 Resumen de Cajas y Picking")

uploaded_file = st.file_uploader("Subí tu archivo Excel (.xlsx):", type=["xlsx"])

if uploaded_file:
    df_raw = pd.read_excel(uploaded_file, sheet_name="LASER", header=None)
    fila_17 = df_raw.iloc[16].fillna("").astype(str).tolist()

    try:
        col_codigo = fila_17.index("CODIGO ADMIN")
        col_cajas = fila_17.index("Cajas")
        col_total_fbo = fila_17.index("Total uds FBO")
    except ValueError:
        st.error("No se encontraron las columnas necesarias en la fila 17.")
        st.stop()

    df = df_raw.iloc[17:, [col_codigo, col_cajas, col_total_fbo]].copy()
    df.columns = ["Codigo", "Cajas", "Total uds FBO"]
    df = df.dropna(subset=["Codigo"])
    df["Cajas"] = pd.to_numeric(df["Cajas"], errors="coerce").fillna(0)
    df["Total uds FBO"] = pd.to_numeric(df["Total uds FBO"], errors="coerce").fillna(0)

    df = df[df["Total uds FBO"] > 0]
    df = df[df["Cajas"] > 0]

    df["Cajas completas"] = (df["Total uds FBO"] // df["Cajas"]).astype(int)
    df["Unidades sobrantes"] = (df["Total uds FBO"] % df["Cajas"]).astype(int)

    def clasificacion(row):
        if row["Unidades sobrantes"] == 0:
            return "Caja completa"
        porcentaje = row["Unidades sobrantes"] / row["Cajas"]
        if porcentaje >= 0.85:
            return "Falta poco"
        else:
            return "Va a picking"

    df["Clasificacion"] = df.apply(clasificacion, axis=1)

    # Obtener valores de la hoja "TOTALES"
    df_totales = pd.read_excel(uploaded_file, sheet_name="TOTALES", header=None)
    total_fbo = int(df_totales.iloc[2, 1])  # fila 3, columna B
    total_def = int(df_totales.iloc[4, 1])  # fila 5, columna B
    total_general = total_fbo + total_def

    st.subheader(":clipboard: Tabla de resumen")
    st.dataframe(df, use_container_width=True)

    st.subheader(":bar_chart: Gráfico de Clasificación")
    total_cajas = df["Cajas completas"].sum()
    total_sobrantes = df["Unidades sobrantes"].sum()
    fig, ax = plt.subplots(figsize=(6, 6))
    valores = [total_cajas, total_sobrantes]
    colores = ["#1f77b4", "#ff7f0e"]
    wedges, texts = ax.pie(valores, labels=None, startangle=90, colors=colores)
    ax.axis("equal")

    leyenda = [
        f"Cajas completas ({int(total_cajas)})",
        f"Unidades sobrantes ({int(total_sobrantes)})",
        "",
        f"🧩 Total uds FBO (fila 3): {total_fbo}",
        f"❌ Uds defectuosas (fila 5): {total_def}",
        f"📦 Total general: {total_general}"
    ]
    ax.legend(wedges, leyenda[:2], title="Distribución", loc="upper left", bbox_to_anchor=(1, 1))
    plt.figtext(1.05, 0.4, "\n".join(leyenda[3:]), fontsize=10, ha="left", va="top", bbox=dict(facecolor='white', edgecolor='black'))
    st.pyplot(fig)

    st.subheader(":chart_with_upwards_trend: Resumen numérico")
    st.metric("Total de Cajas Formadas", int(total_cajas))
    st.metric("Total de Códigos Analizados", len(df))
    st.metric("Total de Piezas Buenas", int(total_fbo))
    st.info(f"📦 Total general: {total_general} (FBO + defectuosas)")

    # Descargar como Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Cajas Picking"
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ref = f"A1:{chr(65 + len(headers) - 1)}{ws.max_row}"
    table = Table(displayName="ResumenPicking", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(output)

    st.download_button(
        label="📥 Descargar resumen en Excel",
        data=output.getvalue(),
        file_name="Resumen_Cajas_Picking_Formato.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
